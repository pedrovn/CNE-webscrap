# Script para hacer webscraping de los datos desagregados a nivel
# de candidatos de las Elecciones Regionales en Venezuela 2021.
# El script trae de la pagina del CNE, las totalizaciones de electores, 
# actas y votos en general para cada uno de los estados, y totalizaciones 
# desagregadas por partido para los votos de cada uno de los candidatos.
# Hecho por Pedro Vicente Navarro. @pedrovn 2021

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook

# Set Selenium 

PATH = "C:/Users/Pedro/Desktop/pyscrapcne/chromedriver_win32/chromedriver.exe"
driver =  webdriver.Chrome(PATH)
driver.get("https://www2.cne.gob.ve/rm2021")

# Esperar que cargue la página del CNE 

try:
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "div.card-header.d-flex.justify-content-end.paddingtb0"))
    )
except:
    print("PAGINA NO CARGA MALDITA SEA EL CNE")
    driver.quit
else:
    print("PAGINA CARGADA CORRECTAMENTE")

# Set main variables, lists and elements

estados = "div > div:nth-child(1) > select > option:nth-child("
filtrar = "div > div.card-body.ng-star-inserted > filter-form > form > div > div:nth-child(6) > button > span"
dropdown = "tr.candidateRow > td:nth-child(2) > i.fa-angle-down"
adentrocand_1 = "table > tbody:nth-child("
adentrocand_2 = ") > tr:nth-child("
nombre_candidato_1 = "candidate-list > div > div.card-body.text-center > table > tbody:nth-child("
nombre_candidato_2 = ") > tr.candidateRow.candidateDetail > td.d-flex.justify-content-start.align-items-center"
final_img = " > td.d-flex.justify-content-center.align-items-center.py-1 > img"
img_del_1 = "https://www2.cne.gob.ve/assets/evento_files_5/partidos/"
img_del_2 = "."
renglones = "tr.candidateDetail.collapse.ng-star-inserted.show"
listlink = []
listlinkimg = []
dataest1 = "/html/body/app-root/app-home/div/div[2]/div/div[1]/div[3]/data-sheet/div[1]/div[2]/div[1]/div[2]/table/tbody/tr"
dataest2 = "/html/body/app-root/app-home/div/div[2]/div/div[1]/div[3]/data-sheet/div[1]/div[2]/div[2]/div[2]/table/tbody/tr"
dataest3 = "/html/body/app-root/app-home/div/div[2]/div/div[1]/div[3]/data-sheet/div[1]/div[2]/div[3]/div[2]/table/tbody/tr"
tit_elec = ('-','ELECTORES')
tit_elec2 = ('-','ACTAS')
tit_elec3 = ('-','VOTOS')

# Diccionario de partidos 

dict_partidos = {
    "AD": "1", "COPEI": "2", "MAS": "5", "MEP": "6", "PCV": "9", "Alianza_para_el_cambio": "219", "VP": "236", "USTED": "246", "PSUV": "269", "PPT": "276", "Pueblo_unido_multietnico_de_Amazonas": "290", "UNT": "460", "UPV": "513", "Podemos": "514", "Movimiento_Regional_de_avanzada": "518", "UNIDOS_POR_MONAGAS": "535", "ASIS": "546", "POR_MI_PUEBLO": "553", "RENA": "632", "Unidad_Soberana": "799", "Dale": "803", "AP": "814", "CAMINA": "846", "Partido_De_accion_mirandino": "866", "Venezuela_Unida": "868", "ConEnzo": "869", "PIO": "877", "Ecologico": "878", "Tiempo_Social_Nacionalista": "894", "ORA": "906", "YIVI": "907", "TUPAMARO": "931", "NUEVO_PACTO": "938", "ORISUR": "1001", "NUVIPA": "1016", "Somos_Venezuela": "1017", "PRO_GUAYANA": "1030", "TINA_ES_I": "1031", "PMI": "1032", "Movimiento_Libertador_Sucre_Despierta": "1034", "Movimiento_renovacion_socialista": "142", "Union_Progreso": "1043", "Puente": "1044", "UPP89": "1047", "El_Cambio": "1048", "Soluciones": "1049", "Cambiemos": "1050", "LAPIZ": "1051", "Primero_Venezuela": "1053", "Prociudadanos": "1054", "Red_de_Aragua": "1056", "Carabobeños_Por_Carabobo": "1057", "LAIN": "1060", "Fuerza_Vecinal": "1062", "Vision_Futuro_Miranda": "1063", "PAZ_PARTIDO_ACCION_ZULIANA": "1064", "MR": "1065", "BR": "1066", "COMPA": "1067", "MinUNIDAD": "1068", "MPV": "1069", "Convergencia": "1071", "Centrados": "1073", "MUD": "1075", "UNDEO": "1076", "Fuerza_Ciudadana": "1077", "MBJ_Bolivar_Joven": "1078", "UNIPRO": "1079", "BARQUISIMETO_ACTIVO": "1082", "Proyecto_el_guacharo": "1083", "Amemos_Tachira": "1087"
}

# Cálculo de estados a seleccionar

nestados = driver.find_elements_by_css_selector("div > div:nth-child(1) > select > option")
nest = len(nestados) - 3
print('Se encontraron ' + str(nest) + ' estados\n')


# Inicio de scraping por estados 

for x in range(3,len(nestados)):

    print('Consultando: ' + nestados[x].text + '(' + str(x) + ')')

    # Seleccionar Estados uno por uno

    nomestado = driver.find_element_by_css_selector("div > div:nth-child(1) > select > option:nth-child(" + str(x+1) + ")")
    print(str(x))
    clickestado = nomestado.click()
    print('Se tocó el estado ' +  nomestado.text)    
    time.sleep(4)

    # Click en filtrar para acceder a pagina del Estado

    driver.find_element_by_css_selector(filtrar).click()
    print ("Se clickeó para filtrar resultados de " + nomestado.text)
    time.sleep(4)

    # Definir documento excel en el que se van a vaciar los datos

    wb = Workbook()
    nombrearchivo = nomestado.text.replace("EDO.","")
    dest_filename = 'ERM2021-{}.xlsx'.format(nombrearchivo)
    delsheet = wb['Sheet']
    wb.remove(delsheet)

    # GET elementos para totalizar actas, electores y votos generales

    dataest_elec = driver.find_elements_by_xpath(dataest1)
    dataest_actas = driver.find_elements_by_xpath(dataest2)
    dataest_votos = driver.find_elements_by_xpath(dataest3)
    data_cand_estado = driver.find_elements_by_css_selector('tr.candidateRow')

    # Crear primera hoja para resultados generales

    wb.create_sheet(title=(nomestado.text + " TOTALES"))
    ws = wb[nomestado.text + " TOTALES"]

    # Escribir datos generales en la primera hoja

    for x in range(len(data_cand_estado)):

        print(data_cand_estado[x].text)
        agre = data_cand_estado[x].text
        ws.append([agre])

    for x in range(len(tit_elec)):
        agr_tit = tit_elec[x]
        ws.append([agr_tit])

    for x in range(len(dataest_elec)):
        print(dataest_elec[x].text)
        agre2 = (dataest_elec[x].text)
        ws.append([agre2])

    for x in range(len(tit_elec2)):
        agr_tit2 = tit_elec2[x]
        ws.append([agr_tit2])

    for x in range(len(dataest_actas)):
        print(dataest_actas[x].text)
        agre2 = (dataest_actas[x].text)
        ws.append([agre2])

    for x in range(len(tit_elec3)):
        agr_tit3 = tit_elec3[x]
        ws.append([agr_tit3])

    for x in range(len(dataest_votos)):
        print(dataest_votos[x].text)
        agre3 = (dataest_votos[x].text)
        ws.append([agre3])


    # Contar dropdowns por cada candidato 

    ndrops = driver.find_elements_by_css_selector(dropdown)   # Aquí se cuantos dropdowns hay en el estado

    # Indexes generales

    index_links = 2
    nameprint = 0
    indexname = 2
    indexclicks = 1
    index_adentrocand = 2
    listanombrepartidos = []

    # Click a cada dropdown recursivamente para ver datos internos del candidato 

    for x in range(len(ndrops)):               # Para cada dropdown de candidato en el estado, hacer lo siguiente

        if ndrops[x].is_displayed():           # Verificar si está visible para clickearlo
            ndrops[x].click()
            print('se clickeo el ' + str(indexclicks) + ' dropdown')

            # Datos internos del candidato

            nrenglones = len(driver.find_elements_by_css_selector(renglones))
            print("Este candidato fue inscrito por " + str(nrenglones) + " partidos")

                # Construye lista de links para datos e imagenes (que están codificadas en formato CNE)

            for x in range(nrenglones):
                
                # Datos

                linkdataden = (adentrocand_1 + str(index_adentrocand) + adentrocand_2 + str(index_links) + ")")

                # Imágenes 

                img_linkdataden = (adentrocand_1 + str(index_adentrocand) + adentrocand_2 + str(index_links) + ")" + final_img)

                # Se le agrega a las listas de cada uno

                listlink.append(linkdataden)    # Enlaces a datos de cada candidato
                listlinkimg.append(img_linkdataden)     # Enlaces a imagenes de cada uno de los partidos (a lo interno del candidato)

                # Se suma 1 al index de links para que el proximo renglon de candidato funcione
                index_links += 1

            if indexname == 2:
                nom_linkdataden = (nombre_candidato_1 + str(indexname) + nombre_candidato_2)
            
            index_adentrocand += 1 

                # Llamada a cada dato interno del candidato

            for x in range(nrenglones):

                dataadentro = driver.find_elements_by_css_selector(listlink[x])
                
                nomcand = driver.find_element_by_css_selector(nom_linkdataden).text

                img = driver.find_element_by_css_selector(listlinkimg[x])

                img_get_src = img.get_attribute('src')
                img_almosthere = img_get_src.replace(img_del_1,"")
                img_final_name = img_almosthere.split(img_del_2, 1)[0]

                # Cambio de imágenes por nombre de partidos
            
                for key, value in dict_partidos.items():
                    if value == img_final_name:
                        img_final_name = key

                # Mostrar nombre del candidato una sola vez y seleccionar hoja en la que se va a escribir

                if nameprint == 0:
                    print(nomcand)
                    nombre_nueva_hoja = nomcand
                    createws = wb.create_sheet(title=nombre_nueva_hoja)
                nameprint += 1

                ws = wb[nombre_nueva_hoja]

                # Escribir datos de cada candidato

                for x in range(len(dataadentro)):
                    print(img_final_name + " " + dataadentro[x].text)
                    agre = (img_final_name + " " + dataadentro[x].text)
                    ws.append([agre])

            # Agrega valores a indexes para llevar la continuidad
            
            dataadentro.clear()
            listlink.clear()
            listlinkimg.clear()
            index_links = 2
            indexname += 1
            nameprint = 0 
            nom_linkdataden = (nombre_candidato_1 + str(indexname) + nombre_candidato_2)
            indexclicks += 1
            
            time.sleep(2)

    wb.save(filename = dest_filename)

# Una vez terminado, cerrar todo. Arrivederci

driver.quit()
